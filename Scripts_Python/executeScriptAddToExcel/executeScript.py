
import subprocess
import conf
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

def extract_numbers(string):
    return re.findall(r'\b\d+\.\d+\b', string)

def execute_wsl_script():

    # Définir la commande et les arguments
    command = [
#        'wsl',
        'python3',
        conf.script,
        conf.option1,
        conf.arg1,
        conf.option2,
        conf.arg2,
        conf.option3,
        conf.arg3
        ]

    try:
        # Exécuter le script Python sous WSL
        result = subprocess.run(command, capture_output=True, text=True)

        print(command)
        
        if result.returncode != 0:
            return
        
        # Analyser la sortie pour trouver la ligne commençant par start_value
        for line in result.stdout.splitlines():
            if line.startswith(conf.start_value):
                lineToReturn = line.strip()
                numberToReturn = extract_numbers(lineToReturn)
        
        # Ajouter la ligne à une cellule d'un fichier Excel
        excel_file = 'output_file.xlsx'
        sheet_name = 'Sheet1'
        
        # Charger le classeur existant ou en créer un nouveau
        try:
            workbook = load_workbook(excel_file)
        except FileNotFoundError:
            from openpyxl import Workbook
            workbook = Workbook()
            workbook.save(excel_file)
            workbook = load_workbook(excel_file)
        
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        
   
        # Charger le fichier Excel ou en créer un nouveau s'il n'existe pas
        try:
            workbook = load_workbook('output_file.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
        
        # Trouver la première cellule vide dans la colonne A
        row = 1
        while sheet.cell(row=row, column=1).value is not None:
            row += 1
        

        # Ajouter la ligne à la première cellule vide pour un string
        #sheet.cell(row=row, column=1, value=stringToReturn)

        # Insérer le nombre dans la cellule
        sheet.cell(row=row, column=1, value=float(numberToReturn[0]))  # Convertir la liste en nombre
      
        
        # Sauvegarder le fichier Excel
        workbook.save('output_file.xlsx')
    
    except Exception as e:
        print(f"Une erreur est survenue : {e}")

for x in range(conf.nbExec):
  print(x)
  execute_wsl_script()